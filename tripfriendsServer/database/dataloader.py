import openpyxl

def showFileContents(file_name='./unknown_area/unknown_place_list.xlsx', sheet_name="전체"):
    # Open the Excel file

    workbook = openpyxl.load_workbook(file_name)

    # Select a sheet within the workbook
    sheet = workbook[sheet_name]

    # Read the data
    for row in sheet.iter_rows(values_only=True):
        print('name: ', row[1],', address: ', row[4])
        print()
    # Close the workbook
    workbook.close()

def UnknownPlaceList(file_name='./unknown_area/unknown_place_list.xlsx', sheet_name="전체"):
    workbook = openpyxl.load_workbook(file_name)

    # Select a sheet within the workbook
    sheet = workbook[sheet_name]

    place_list = []
    # Read the data
    for row in sheet.iter_rows(values_only=True):
        temp_list = []
        temp_list.append(row[1])
        temp_list.append(row[4])
        place_list.append(temp_list)
    print(place_list)
    # Close the workbook
    workbook.close()
    return place_list

if __name__ == "__main__":
    UnknownPlaceList()